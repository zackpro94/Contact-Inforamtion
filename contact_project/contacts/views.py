from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Q
from .models import Employee
from .forms import EmployeeForm
import openpyxl
from openpyxl import Workbook
import csv

def is_admin(user):
    return user.is_authenticated and user.is_staff

@login_required
@user_passes_test(is_admin)
def home(request):
    search_query = request.GET.get('search', '')
    employees = Employee.objects.filter(admin=request.user).order_by('-created_at')
    
    if search_query:
        employees = employees.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(company_name__icontains=search_query)
        )
    
    paginator = Paginator(employees, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'contacts/home.html', {'page_obj': page_obj, 'search_query': search_query})

@login_required
@user_passes_test(is_admin)
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.admin = request.user  # Assign the current admin
            employee.save()
            return redirect('home')
    else:
        form = EmployeeForm()
    return render(request, 'contacts/add_employee.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id, admin=request.user)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'contacts/edit_employee.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id, admin=request.user)
    employee.delete()
    return redirect('home')


def view_employee(request, slug):
    employee = get_object_or_404(Employee, slug=slug)
    return render(request, 'contacts/view_employee.html', {'employee': employee})


def export_employees(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employees.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Full Name', 'Email', 'Phone', 'Company', 'Website', 'Address'])
    
    employees = Employee.objects.filter(admin=request.user)
    for emp in employees:
        writer.writerow([emp.full_name, emp.email, emp.phone_number, 
                        emp.company_name, emp.company_website, emp.address])
    
    return response

def export_employees_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="employees.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = ['Full Name', 'Email', 'Phone', 'Company', 'Website', 'Address']
    ws.append(headers)
    
    employees = Employee.objects.filter(admin=request.user)
    for emp in employees:
        ws.append([emp.full_name, emp.email, emp.phone_number, 
                 emp.company_name, emp.company_website, emp.address])
    
    wb.save(response)
    return response

def import_employees(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2):
            Employee.objects.create(
                admin=request.user,
                full_name=row[0].value,
                email=row[1].value,
                phone_number=row[2].value,
                company_name=row[3].value,
                company_website=row[4].value,
                address=row[5].value
            )
        
        return redirect('home')
    return render(request, 'contacts/import.html')